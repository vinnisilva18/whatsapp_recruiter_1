from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from whatsapp_recruiter.config import Settings
from whatsapp_recruiter.models import CandidateRecord


class CandidateExporter:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    def export(self, records: list[CandidateRecord], output_base: Path) -> tuple[Path, Path]:
        json_path = output_base.with_suffix(".json")
        xlsx_path = output_base.with_suffix(".xlsx")

        data = [record.as_export_dict() for record in records]
        frame = pd.DataFrame(data)
        frame.to_json(json_path, orient="records", indent=2, force_ascii=False)
        self._export_xlsx(records, xlsx_path)
        return json_path, xlsx_path

    def _export_xlsx(self, records: list[CandidateRecord], xlsx_path: Path) -> None:
        workbook = self._build_workbook(xlsx_path)
        corpo_sheet = self._get_or_create_sheet(
            workbook,
            "Corpo Clínico",
            [
                "CRM/SP",
                "Nome ",
                "CPF ",
                "Email ",
                "Celular ",
                "Validade Cadastro SAM ",
                "Data de Nascimento",
                "Especialidade ",
                "PJ",
                "Observações ",
            ],
        )
        processo_sheet = self._get_or_create_sheet(
            workbook,
            "Cadastros - Em processo ",
            [
                "CRM",
                "Nome ",
                "Celular ",
                "E-mail ",
                "Indicação ",
                "Especialidade",
                "Visita PS ",
                "Carta de Apresentação ",
                "Status SAM ",
                "OBS ",
            ],
        )

        self._clear_data_rows(corpo_sheet)
        self._clear_data_rows(processo_sheet)

        for record in records:
            corpo_sheet.append(
                [
                    self._crm_display(record),
                    record.nome_completo,
                    record.cpf,
                    record.email,
                    record.telefone,
                    None,
                    record.data_nascimento,
                    record.especializacao,
                    None,
                    self._build_observacoes(record),
                ]
            )
            processo_sheet.append(
                [
                    self._crm_display(record),
                    record.nome_completo,
                    record.telefone,
                    record.email,
                    record.indicacao,
                    record.especializacao,
                    None,
                    record.carta_apresentacao,
                    record.status_sam,
                    self._build_observacoes(record),
                ]
            )

        workbook.save(xlsx_path)

    def _build_workbook(self, xlsx_path: Path):
        template_path = self.settings.template_xlsx_path
        if template_path and template_path.exists():
            shutil.copyfile(template_path, xlsx_path)
            return load_workbook(xlsx_path)
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        return workbook

    @staticmethod
    def _get_or_create_sheet(workbook, sheet_name: str, headers: list[str]):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        return sheet

    @staticmethod
    def _clear_data_rows(sheet) -> None:
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)

    @staticmethod
    def _crm_display(record: CandidateRecord) -> str | None:
        if not record.crm:
            return None
        if record.crm_uf:
            return f"{record.crm}/{record.crm_uf}"
        return record.crm

    @staticmethod
    def _build_observacoes(record: CandidateRecord) -> str | None:
        parts = []
        if record.cidade and record.uf:
            parts.append(f"Localidade: {record.cidade}/{record.uf}")
        return " | ".join(parts) if parts else record.observacoes
